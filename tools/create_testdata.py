"""
Erstellt den M17 Standard-Testdatensatz für Testautomatisierung.
Enthält gezielte Edge Cases aus TestAutomatisierung.md Block B.
"""
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import os

def d(year, month, day):
    """Hilfsfunktion: datetime Objekt"""
    return datetime(year, month, day)

wb = Workbook()

# ══════════════════════════════════════════════════════════════
# Sheet 1: JiraStories (Pflicht-Sheet)
# ══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'JiraStories'

headers = [
    'Jira-ID', 'Issue-Type', 'Squad', 'Issue-Status',
    'Ready4Progress_first', 'Ready4Progress',
    'In Progress_first', 'In Progress',
    'leaving_In Progress_first', 'leaving_In Progress',
    'Ready4Test_first', 'Ready4Test',
    'leaving_Ready4Test_first', 'leaving_Ready4Test',
    'Resolved', 'Rejected',
]
ws.append(headers)

# ── Normalfall: abgeschlossen, Squad A ──────────────────────
# CT = In Progress bis Resolved
ws.append(['JA-001', 'Story', 'Squad-A', 'Resolved',
    d(2024,1,1), d(2024,1,1),    # Ready4Progress_first == Ready4Progress (einmal)
    d(2024,1,2), d(2024,1,2),    # In Progress_first == In Progress (einmal)
    d(2024,1,8), d(2024,1,8),    # leaving_In Progress_first == leaving
    d(2024,1,9), d(2024,1,9),    # Ready4Test_first == Ready4Test
    d(2024,1,12), d(2024,1,12),  # leaving_Ready4Test
    d(2024,1,15), None           # Resolved, kein Rejected
])

# ── Dual-Period: Status zweimal durchlaufen ──────────────────
# In Progress: 3.1–5.1 (3 Tage) + 10.1–11.1 (2 Tage) = 5 Tage
ws.append(['JA-002', 'Story', 'Squad-A', 'Resolved',
    d(2024,1,3), d(2024,1,3),
    d(2024,1,3), d(2024,1,10),   # _first = 3.1, Basis = 10.1 → zweimal
    d(2024,1,5), d(2024,1,11),   # leaving: _first endet 5.1, Basis endet 11.1
    d(2024,1,12), d(2024,1,12),
    d(2024,1,14), d(2024,1,14),
    d(2024,1,20), None
])

# ── Aktives WIP-Item (weder Resolved noch Rejected) ─────────
ws.append(['JA-003', 'Bug', 'Squad-A', 'In Progress',
    d(2024,2,1), d(2024,2,1),
    d(2024,2,2), d(2024,2,2),
    None, None, None, None, None, None,
    None, None
])

# ── Rejected Item (abgebrochen, nicht in Rolling Pace) ──────
ws.append(['JA-004', 'Story', 'Squad-B', 'Rejected',
    d(2024,1,5), d(2024,1,5),
    d(2024,1,6), d(2024,1,6),
    d(2024,1,10), d(2024,1,10),
    None, None, None, None,
    None, d(2024,1,12)
])

# ── Item ohne Squad (leere Squad-Spalte) ────────────────────
ws.append(['JA-005', 'Task', '', 'Resolved',
    d(2024,1,8), d(2024,1,8),
    d(2024,1,9), d(2024,1,9),
    d(2024,1,13), d(2024,1,13),
    d(2024,1,14), d(2024,1,14),
    d(2024,1,14), d(2024,1,17), None
])

# ── Squad-B: genau 1 Item (Grenzfall für Statistiken) ───────
ws.append(['JB-001', 'Story', 'Squad-B', 'Resolved',
    d(2024,3,1), d(2024,3,1),
    d(2024,3,2), d(2024,3,2),
    d(2024,3,7), d(2024,3,7),
    d(2024,3,8), d(2024,3,8),
    d(2024,3,10), None
])

# ── Squad-C: 20 Items für Normalfall (CT variiert) ──────────
for i in range(20):
    start = d(2024, 1, 1) + timedelta(days=i * 3)
    ct = 3 + (i % 7)  # CT variiert zwischen 3 und 9 Tagen
    ws.append([
        f'JC-{i+1:03d}', 'Story', 'Squad-C', 'Resolved',
        start, start,
        start + timedelta(days=1), start + timedelta(days=1),
        start + timedelta(days=ct), start + timedelta(days=ct),
        start + timedelta(days=ct+1), start + timedelta(days=ct+1),
        start + timedelta(days=ct+2), start + timedelta(days=ct+2),
        start + timedelta(days=ct+2), None
    ])

# ── Inkonsistentes Datum: Ende vor Start (Datenqualitätsproblem) ──
ws.append(['JA-ERR', 'Story', 'Squad-A', 'Resolved',
    d(2024,4,10), d(2024,4,10),
    d(2024,4,15), d(2024,4,15),   # In Progress startet später als Resolved
    d(2024,4,12), d(2024,4,12),   # leaving_In Progress VOR In Progress → negativ
    None, None, None, None,
    d(2024,4,10), None            # Resolved = gleicher Tag wie Start
])

# ══════════════════════════════════════════════════════════════
# Sheet 2: Epics (für Say_Do_Ratio)
# ══════════════════════════════════════════════════════════════
ws_epics = wb.create_sheet('Epics')
ws_epics.append(['Epic-ID', 'Summary', 'Squad', 'Iteration', 'Resolved', 'Rejected'])
ws_epics.append(['EP-001', 'Feature A', 'Squad-A', 'Q1-2024', d(2024,3,28), None])
ws_epics.append(['EP-002', 'Feature B', 'Squad-A', 'Q1-2024', None, None])   # nicht geliefert
ws_epics.append(['EP-003', 'Feature C', 'Squad-B', 'Q1-2024', d(2024,3,30), None])
ws_epics.append(['EP-004', 'Feature D', 'Squad-A', 'Q2-2024', d(2024,6,28), None])
ws_epics.append(['EP-005', 'Feature E', 'Squad-A', 'Q2-2024', d(2024,6,29), None])
ws_epics.append(['EP-006', 'Feature F', 'Squad-B', 'Q2-2024', None, d(2024,5,15)])  # rejected

# ══════════════════════════════════════════════════════════════
# Sheet 3: Happiness Faktor (Custom-Header in Zeile 3)
# ══════════════════════════════════════════════════════════════
ws_hf = wb.create_sheet('Happiness Faktor')
ws_hf.append(['Flow Analytics – Happiness Export'])          # Zeile 1: Titel
ws_hf.append(['Stand: 2024-06-01'])                          # Zeile 2: Meta
ws_hf.append(['Schlüsselwert', 'Squad', 'Jan 2024', 'Feb 2024', 'Mrz 2024', 'Apr 2024'])  # Zeile 3: Header
ws_hf.append([1, 'Squad-A', 4, 3, 4, 5])
ws_hf.append([2, 'Squad-B', 3, 3, 2, 4])
ws_hf.append([3, 'Squad-C', 5, 4, 4, 3])

# ══════════════════════════════════════════════════════════════
# Speichern
# ══════════════════════════════════════════════════════════════
out_dir = os.path.join(os.path.dirname(__file__), '..', 'tests', 'fixtures')
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, 'testdata.xlsx')
wb.save(out_path)
print(f'OK testdata.xlsx: {os.path.abspath(out_path)}')

# ── Leere Datei (nur Header, keine Datenzeilen) ──────────────
wb_empty = Workbook()
ws_empty = wb_empty.active
ws_empty.title = 'JiraStories'
ws_empty.append(headers)
empty_path = os.path.join(out_dir, 'testdata-empty.xlsx')
wb_empty.save(empty_path)
print(f'OK testdata-empty.xlsx: {os.path.abspath(empty_path)}')

# ── Einzelner-Squad-Testdatensatz (testdata-single-squad.xlsx) ────────────
# Testet den Sonderfall: nur 1 Squad → Auto-Setzung des Squad-Filters
wb_single = Workbook()
ws_sq = wb_single.active
ws_sq.title = 'JiraStories'
ws_sq.append(headers)

# 5 abgeschlossene Items für Squad-X
for i in range(5):
    start = d(2024, 3, 1) + timedelta(days=i * 4)
    ct = 4 + i
    ws_sq.append([
        f'SX-{i+1:03d}', 'Story', 'Squad-X', 'Resolved',
        start, start,
        start + timedelta(days=1), start + timedelta(days=1),
        start + timedelta(days=ct), start + timedelta(days=ct),
        start + timedelta(days=ct+1), start + timedelta(days=ct+1),
        start + timedelta(days=ct+2), start + timedelta(days=ct+2),
        start + timedelta(days=ct+2), None
    ])

# 1 aktives WIP-Item (für WIP-Tile)
ws_sq.append(['SX-006', 'Bug', 'Squad-X', 'In Progress',
    d(2024, 5, 1), d(2024, 5, 1),
    d(2024, 5, 2), d(2024, 5, 2),
    None, None, None, None, None, None,
    None, None
])

# Happiness Faktor Sheet (nur Squad-X)
ws_hf_sq = wb_single.create_sheet('Happiness Faktor')
ws_hf_sq.append(['Flow Analytics – Happiness Export'])
ws_hf_sq.append(['Stand: 2024-06-01'])
ws_hf_sq.append(['Schlüsselwert', 'Squad', 'Jan 2024', 'Feb 2024', 'Mrz 2024', 'Apr 2024'])
ws_hf_sq.append([1, 'Squad-X', 4, 3, 4, 5])

single_path = os.path.join(out_dir, 'testdata-single-squad.xlsx')
wb_single.save(single_path)
print(f'OK testdata-single-squad.xlsx: {os.path.abspath(single_path)}')
